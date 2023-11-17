class parser:
    def __init__(self, boolpic=True, DownloadVar=None, SaveBD=True):
        self._LINK = "https://satisfactory-calculator.com"
        self._ID = 0
        self._boolpic = boolpic

        self.ITEMS = []
        self.STRUCTURS = []
        self.CRAFTS = []

        self.LINKS = []
        self.sITEMS = {}
        self.sSTRUCTURS = {}

        if self._TryConnect(self._LINK):
            self._ParserItems(DownloadVar)
            self._ParserStructure(DownloadVar)
            DownloadVar.set(0.7)
            self._ParserCraft(DownloadVar)
            if SaveBD:
                self._importEx()
            else:
                print(f"MAIN DATA:\nITEMS: {len(self.ITEMS)} {self.ITEMS}\nSTRUCTURS: {len(self.STRUCTURS)} {self.STRUCTURS}\nCRAFTS: {len(self.CRAFTS)} {self.CRAFTS}\nSIDE DATA:\nLINKS: {len(self.LINKS)} {self.LINKS}\nSET ITEMS: {len(self.sITEMS.items())} {self.sITEMS.items()}\nSET STRUCTURS: {len(self.sSTRUCTURS.items())} {self.sSTRUCTURS.items()}\n")
            DownloadVar.set(1.00)
        else:
            print("Lost Connecting")
    def _TryConnect(self, link):
        import requests
        return requests.get(link).status_code==200

    def _ParserItems(self, DownloadVar):
        print("ParserItemsInit")
        import requests
        from os import mkdir
        from bs4 import BeautifulSoup as bs
        site = requests.get(self._LINK + "/en/items")
        soup = bs(site.text, 'html.parser')
        col6 = soup.findAll('div', class_='container-fluid')[1]
        if self._boolpic:
            try:
                mkdir("pic")
                mkdir("TEMP")
            except:
                pass
        TEMP = None
        for i, ic in enumerate(col6):
            if (i) % 3 == 0 and i % 2 != 0:
                for link in ic.find_all('img'):
                    self._ID += 1
                    src = link.get('src')
                    if self._boolpic:
                        with open(f"pic/{self._ID}.png","wb") as file:
                            file.write(requests.get(src).content)
                        print(f"Download pic: {self._ID} {src}")
                        DownloadVar.set(0.004*self._ID)
                    alt = link.get("alt")
                    if TEMP=="Ores" or TEMP=="Minerals" or TEMP=="Liquids" or TEMP=="Gas" or TEMP=="Fuels":
                        self.ITEMS.append((self._ID, alt, TEMP))
                    else:
                        self.ITEMS.append((self._ID, alt))
                    self.sITEMS[alt] = self._ID
                nt = -1
                n = -1
                for link in ic.find_all('a'):
                    n += 1
                    if nt < 15 and n % 2 == 0:
                        if nt != 4:
                            href = self._LINK + link.get('href')
                            self.LINKS.append(href)
                        nt += 1
        print("ParserItemsDone")

    def _ParserStructure(self, DownloadVar):
        print("ParserStructureInit")
        import requests
        from bs4 import BeautifulSoup as bs
        site = requests.get(self._LINK + "/en/buildings")
        soup = bs(site.text, 'html.parser')
        col6 = soup.findAll('div', class_='container-fluid')[1]
        for link in col6.find_all('img'):
            self._ID += 1
            src = link.get('src')  # pic
            if self._boolpic:
                with open(f"pic/{self._ID}.png", "wb") as file:
                    file.write(requests.get(src).content)
                print(f"Download pic: {self._ID} {src}")
                DownloadVar.set(0.004 * self._ID)
            alt = link.get("alt")  # name
            self.STRUCTURS.append((self._ID, alt))
            self.sSTRUCTURS[alt] = self._ID
        print("ParserStructureDone")

    def _ParserCraft(self, DownloadVar):
        def RETURNINT(text):
            num = ("0","1","2","3","4","5","6","7","8","9")
            x = ""
            for i in text:
                if i in num:
                    x+=i
            return int(x)
        print("ParserCraftInit")
        for l, lc in enumerate(self.LINKS):
            DownloadVar.set(1/(117*3.31)*l+0.7)
            print(f"link {l} {lc}")
            import requests
            from bs4 import BeautifulSoup as bs
            site = requests.get(lc)
            soup = bs(site.text, 'html.parser')
            name_ = soup.find("h4").text
            Recipes = soup.findAll("div", class_="card h-100")
            R = Recipes[3].find("strong")
            if R is not None:
                if R.text=="Recipes":
                    d1, d2, ERR = [],[], False
                    R2= Recipes[3].find_all_next("div", class_="row align-items-center")
                    name = self.sSTRUCTURS[R2[1].a.text]
                    TEMP = [[None, None, None, None], [None, None, None, None], [None, None], [None, None]]
                    for i in R2[2].findNext("div").strings:
                        d1.append(i)
                    for i in R2[2].div.next_sibling.next_sibling.strings:
                        d2.append(i)
                    for i in range(len(d1)//4):
                        ti = i*4
                        if d1[ti+3] in self.sITEMS:
                            TEMP[0][i] = self.sITEMS[d1[ti + 3]]
                        else:
                            ERR = True
                        TEMP[1][i] = RETURNINT(d1[ti])
                    for i in range(len(d2)//4):
                        ti = i*4
                        try:
                            if d2[ti+3] == name_:
                                TEMP[3][0] = self.sITEMS[d2[ti+3]]
                        except:
                            ERR = True
                        TEMP[2][0] = RETURNINT(d2[ti])
                    if not(ERR) and not(name_ in ("Silica","Aluminum Scrap","Alien Protein","Water","Crude Oil","Fuel","Sulfuric Acid","Nitrogen Gas")):
                        self.CRAFTS.append((name, TEMP[0][0], TEMP[0][1], TEMP[0][2], TEMP[0][3], TEMP[1][0], TEMP[1][1],
                                           TEMP[1][2], TEMP[1][3], TEMP[2][0], TEMP[3][0]))
                else:
                    continue

    def _importEx(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Items"
        ws2 = wb.create_sheet('Structures')
        ws3 = wb.create_sheet('Crafts')

        for row in self.ITEMS:
            ws1.append(row)
        for row in self.STRUCTURS:
            ws2.append(row)
        for row in self.CRAFTS:
            ws3.append(row)

        wb.save('BD.xlsx')

if __name__ == '__main__':
    class DownloadVarPass:
        def __init__(self):
            pass
        def set(self, x):
            pass
    parser(False, DownloadVarPass(), False)